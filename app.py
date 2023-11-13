import json
from pprint import pprint
import shutil
import sys
from time import time
import tkinter as tk
from tkinter import messagebox as mb

import pandas as pd
from openpyxl import load_workbook

from checkers import check_addres, check_city, check_side, check_sum
from files import clean_output, get_file, check_folders
from settings import FROM_PATH, OUTPUT_PATH, SOURCE_HEADERS_INDEX, TO_PATH


def process(window: tk.Tk, cities):
    try:
        window.destroy()
        listnames = cities

        from_path = FROM_PATH
        to_path = TO_PATH
        output_path = OUTPUT_PATH
        logfile_path = output_path + "log.txt"
        result_path = output_path + "result.xlsx"
        from_file = get_file(from_path)
        to_file = get_file(to_path)

        check_folders(from_path, to_path, output_path)
        clean_output(output_path)

        source_df = pd.read_excel(from_file, sheet_name=1, header=SOURCE_HEADERS_INDEX)
        source_dicts = source_df.to_dict(orient="records")
        print(f"Всего в тендерных ценах: {len(source_dicts)}")

        list_counter = 0
        results = {}
        not_matched = []

        while list_counter < len(listnames):
            try:
                list_df = pd.read_excel(to_file, sheet_name=listnames[list_counter])
            except Exception as e:
                print(
                    f"Невозможно открыть страницу {listnames[list_counter]}, проверьте корректность названия"
                )
                sys.exit()
            list_city = None
            for item in list_df.to_dict(orient="records"):
                if list_city is None:
                    list_city = item["Город"]
                    print(f"Загружаем город: {list_city}")
                    city_source_dicts = list(
                        filter(lambda x: x["Город"] == list_city[:-1], source_dicts)
                    )
                    print(f"В городе значений: {len(city_source_dicts)}")
                source_item = find_item(item, city_source_dicts)
                if source_item is None:
                    not_matched.append(item["Адрес"])
                    continue
                if not check_sum(item, source_item):
                    results[item["#"]] = source_item["Цена ролик 5 сек"]
            list_counter += 1

        shutil.copy2(to_file, result_path)
        pprint(results)
        mb.showinfo(title='Процесс начат',message="Началось создание нового файла, вы получите информацию по завершению")
        workbook = load_workbook(result_path)
        for city in listnames:
            workbook_list = workbook[city]
            price_column = None
            for index, cell in enumerate(workbook_list[1], start=1):
                if cell.value == "Итого, NET":
                    price_column = index
                    break
            if price_column is None:
                raise ValueError(f"Не найдена колонка на странице {city} - Итого, NET")
            counter = 1
            for row in workbook_list.iter_rows(
                min_row=2, max_row=workbook_list.max_row, values_only=True
            ):
                counter += 1
                if row[0] in results.keys():
                    cell = workbook_list.cell(counter, price_column)
                    cell.value = results[row[0]]
        workbook.save(result_path)
        workbook.close()
        with open(logfile_path, "w", encoding="utf-8") as file:
            json.dump(results, file, indent=4, ensure_ascii=False)
        mb.showinfo(title='Процесс завершен', message="Успешно")
    except Exception as e:
        print(e)


def create_tkinter_app(window: tk.Tk):
    frame = tk.Frame(window, padx=10, pady=10)
    frame.pack(expand=True)

    height_lb = tk.Label(
        frame,
        text="Введите через запятую без \n пробелов название \n листов, которые хотите проверить",
    )
    height_lb.grid(row=3, column=1)
    list_input = tk.Entry(
        frame,
    )
    list_input.grid(row=4, column=2, pady=5)
    cal_btn = tk.Button(
        frame,
        text="Подтвердить",
        command=lambda: process(window, str(list_input.get()).split(","))
    )
    cal_btn.grid(row=5, column=2)
    


def find_item(item, source_dicts) -> dict | None:
    finded_item = None
    filtered_sources = list(
        filter(
            lambda x: check_city(item["Город"], x["Город"])
            and check_side(item["Сторона"], x["Сторона"]),
            source_dicts,
        )
    )
    filtered_by_address = list(
        filter(lambda x: check_addres(x["Адрес"], item["Адрес"]), filtered_sources)
    )
    if filtered_by_address:
        finded_item = filtered_by_address[0]
        if len(filtered_by_address) > 1:
            print(f'For addrees {item["Адрес"]} finded many {filtered_by_address}')
    return finded_item


if __name__ == "__main__":
    window = tk.Tk()
    window.title("Проверка заполнения файлов")
    window.geometry("800x600")
    t = time()
    create_tkinter_app(window)
    window.mainloop()
    print("Время работы программы:", round(time() - t, 2))

# [x] to - шаблон
# [x] from - тендерные цены
# [ ] если нет колонок в шаблоне - город-адрес-сторона-цена,NET - вывести ошибки
# [x] в виде скрипта с двумя папками - шаблоны и тендерные цены
# [x] лог в файлик log%date%-%timestamp%.txt
# [x] gui на tkinter
